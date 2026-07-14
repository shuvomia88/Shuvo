import os
import json
import re
import threading
import datetime
import uuid
import random
import pyotp
import logging
import asyncio
import requests
from http.server import BaseHTTPRequestHandler, HTTPServer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters
from telegram.error import TelegramError, Conflict

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s"
)
logger = logging.getLogger(__name__)

# httpx প্রতিটা Telegram API কল-এ একটা INFO লাইন প্রিন্ট করে (getUpdates,
# getMe ইত্যাদি) — এতে console অকারণে ভরে যায়, তাই এটার লেভেল WARNING-এ
# তুলে দেওয়া হলো যাতে শুধু আসল সমস্যা হলেই কিছু দেখায়।
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("apscheduler").setLevel(logging.WARNING)

# ============================================================
# CONFIG & FILE SETTINGS
# ============================================================

BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", "6470499890"))

# বাধ্যতামূলক চ্যানেলগুলোর ইউজারনেম (বটকে অবশ্যই এই চ্যানেলে এডমিন হতে হবে)
REQUIRED_CHANNELS = ["@range_channele", "@insagramth"]

_lock = threading.Lock()

def _style(btn, style_name: str):
    """
    কিছু বাটনে (রঙিন UI hint-এর জন্য) 'style' নামে একটা কাস্টম attribute
    বসানোর চেষ্টা করা হয়। নতুন python-telegram-bot ভার্সনে Keyboard/
    InlineKeyboardButton ক্লাস __slots__ ব্যবহার করে বলে সরাসরি নতুন
    attribute বসানো সবসময় সম্ভব হয় না — এই ফাংশন চেষ্টা করে, ব্যর্থ হলে
    নিঃশব্দে স্কিপ করে (বাটন ঠিকই কাজ করবে, শুধু রঙের hint-টা যোগ হবে না)।
    """
    try:
        object.__setattr__(btn, 'style', style_name)
    except Exception:
        pass

# ============================================================
# FIREBASE REALTIME DATABASE (সব ডেটা এখানেই স্থায়ীভাবে সেভ হবে)
# Render-এ ডিস্ক ephemeral, তাই local JSON file এর বদলে Firebase
# ব্যবহার করা হচ্ছে যাতে redeploy/restart এ ডেটা হারিয়ে না যায়।
# ============================================================
FIREBASE_BASE_URL = "https://realtime-database-7310e-default-rtdb.firebaseio.com"
_DATA_CACHE = None  # in-memory cache, _lock দ্বারা সুরক্ষিত

def _firebase_get(path=""):
    """Firebase থেকে ডেটা রিড করার ফাংশন"""
    try:
        r = requests.get(f"{FIREBASE_BASE_URL}/{path}.json", timeout=10)
        if r.status_code == 200:
            return r.json()
        logger.error(f"Firebase GET failed on '{path}': HTTP {r.status_code}")
    except Exception as e:
        logger.error(f"Firebase GET Error on '{path}': {e}")
    return None

def _firebase_patch(data: dict, path=""):
    """Firebase-এ শুধু নির্দিষ্ট top-level key গুলো আপডেট করে (অন্য কোনো ডেটা মোছে না)"""
    try:
        r = requests.patch(f"{FIREBASE_BASE_URL}/{path}.json", json=data, timeout=10)
        if r.status_code == 200:
            return True
        logger.error(f"Firebase PATCH failed on '{path}': HTTP {r.status_code}")
    except Exception as e:
        logger.error(f"Firebase PATCH Error on '{path}': {e}")
    return False

def _firebase_put(data, path=""):
    """Firebase-এ একটা নির্দিষ্ট path সম্পূর্ণভাবে overwrite করে (delete সহ ঠিকভাবে reflect হয়)"""
    try:
        r = requests.put(f"{FIREBASE_BASE_URL}/{path}.json", json=data, timeout=10)
        if r.status_code == 200:
            return True
        logger.error(f"Firebase PUT failed on '{path}': HTTP {r.status_code}")
    except Exception as e:
        logger.error(f"Firebase PUT Error on '{path}': {e}")
    return False

# ============================================================
# ID DUMP (Facebook/Instagram সাবমিশন এখানে জমা হবে, Admin Panel
# থেকে txt ফাইল হিসেবে ডাউনলোড করলে খালি হয়ে যাবে)
# ============================================================
def add_to_id_dump(category: str, entry: dict):
    """
    একটা সাবমিশনের তথ্য (task, username, password, twofa) Facebook/Instagram
    ডাম্প লিস্টে structured আকারে যোগ করে (Firebase-এ persist থাকে), যাতে
    পরে টাস্ক অনুযায়ী গ্রুপ করে সাজানো যায়। কোনো কারণে ব্যর্থ হলে চুপচাপ
    log করে, বট থামে না।
    """
    key = "fb_dump" if category == "facebook" else "ig_dump"
    try:
        with _lock:
            d = _load()
            if key not in d:
                d[key] = []
            d[key].append(entry)
            _save(d)
    except Exception as e:
        logger.error(f"ID ডাম্পে যোগ করা যায়নি ({category}): {e}")

# ============================================================
# MULTI-LANGUAGE DICTIONARY
# ============================================================
LANGUAGES = {
    "en": {
        "welcome": "🎉 Welcome to {name} TaskPro Bot! 🤖\n\nComplete tasks, submit your proof, and earn rewards with ease.\n\n🚀 Happy Earning & Good Luck!",
        "btn_balance": "💳 BALANCE",
        "btn_tasks": "📋 TASKS",
        "btn_withdraw": "📤 WITHDRAW",
        "btn_report": "📊 YOUR REPORT",
        "btn_support": "ℹ️ SUPPORT",
        "btn_language": "🌐 LANGUAGE",
        "btn_admin": "🛠️ ADMIN PANEL",
        "btn_back": "🔙 BACK",
        "btn_cancel": "❌ cancel",
        "btn_start": "▶️ Start",
        "btn_video": "🎥 Video",
        "btn_how_to_2fa": "❓ How to get 2fa?",
        "select_lang": "🌐 Select Language / ভাষা নির্বাচন করুন:",
        "lang_changed": "✅ Language changed to English!",
        "balance_msg": "💳 Your Balance: ${bal}",
        "report_msg": "📊 All Account Report\n\n✅ Success: [{s}]\n⏳ Reviewing: [{r}]\n❌ Rejected: [{rej}]",
        "select_cat": "📋 Select Category:",
        "task_hidden": "❌ This task is currently hidden by Admin.",
        "no_tasks": "❌ No tasks available in this category.",
        "choose_type": "🎯 Choose Task:",
        "send_2fa_secret": "👉 Please Send Your 2FA Secret Key",
        "send_cookies": "👉 Please Send Your Cookies Data",
        "send_fb_uid": "👉 Please Send Your Facebook UID",
        "invalid_2fa": "❌ Invalid 2FA Secret Key! Please send a valid key again:",
        "withdraw_dash": "💳 Your Balance Dashboard\n\n💰 Balance: ${bal}\n💸 Minimum Withdraw: $0.20\n💳 Withdrawal Charge: $0.03\n✅ You Will Receive: ${rec}",
        "withdraw_min_err": "❌ Unsuccessful balance: Minimum $0.20 required",
        "select_meth": "💳 Select Your Withdraw Method:",
        "send_num": "📱 Please Send Your {method} Number",
        "enter_amt": "💰 Please Enter Your Withdraw Amount:",
        "insufficient": "❌ Insufficient balance.",
        "min_amt_err": "❌ Minimum withdraw is $0.20",
        "check_info": "📋 Please Check Your Info\n\n📱 Number: {num}\n💳 Method: {method}\n💵 Amount: ${amt}\n✅ Receive: ${rec}\n\n✅ If All Information Is Correct, Please Tap the Confirm Button. 👇",
        "btn_confirm": "✅ Confirm",
        "pay_pending": "⏳ Your Payment Is Pending.\n\n👨‍💼 Please Wait for Admin Approval.",
        "cookies_rec": "👉 Cookies Received. Click below to proceed.",
        "btn_acc_reg": "✅ Account Registered",
        "invite_check": "⚠️ Have You Invited 2 Friends?\n\n❌ If You Have Not Invited 2 Friends, Your Report Will Be Rejected.",
        "invite_check_short": "✅ Please confirm below to submit your report.",
        "btn_subbed": "✅ Yes | I Am Subscribed",
        "thanks_msg": "✅ Thanks! Please Do Not Unfollow. Follow the Rules.",
        "report_received": "✅ Your report has been received!\n⏳ Please wait 16–24 hours.",
        "no_usernames_err": "❌ No user available yet!",
        "force_join_msg": "📢 আমাদের বটটি ব্যবহার করতে নিচের চ্যানেলগুলোতে জয়েন করুন:",
        "not_joined_all": "❌ আপনি এখনো সবগুলো চ্যানেলে জয়েন করেননি! দয়া করে জয়েন করে আবার ভেরিফাই করুন।",
        "verify_success": "✅ এখন আমাদের বটটি ব্যবহার করতে পারবেন।\nআমাদের চ্যানেলে জয়েন হওয়ার জন্য ধন্যবাদ! ❤️",
        "support_msg": "⚠️ কোনো সমস্যা হলে অ্যাডমিনকে জানান।"
    },
    "bn": {
        "welcome": "🎉 Welcome to {name} TaskPro Bot! 🤖\n\nComplete tasks, submit your proof, and earn rewards with ease.\n\n🚀 Happy Earning & Good Luck!",
        "btn_balance": "💳 ব্যালেন্স",
        "btn_tasks": "📋 কাজ (TASKS)",
        "btn_withdraw": "📤 টাকা তুলুন",
        "btn_report": "📊 আপনার রিপোর্ট",
        "btn_support": "ℹ️ সাপোর্ট (SUPPORT)",
        "btn_language": "🌐 ভাষা (LANGUAGE)",
        "btn_admin": "🛠️ এডমিন প্যানেল",
        "btn_back": "🔙 পেছনে যান",
        "btn_cancel": "❌ বাতিল করুন",
        "btn_start": "▶️ শুরু করুন",
        "btn_video": "🎥 ভিডিও দেখুন",
        "btn_how_to_2fa": "❓ How to get 2fa?",
        "select_lang": "🌐 Select Language / ভাষা নির্বাচন করুন:",
        "lang_changed": "✅ ভাষা পরিবর্তন করে বাংলায় সেট করা হয়েছে!",
        "balance_msg": "💳 আপনার বর্তমান ব্যালেন্স: ${bal}",
        "report_msg": "📊 সকল অ্যাকাউন্ট রিপোর্ট\n\n✅ সফল: [{s}]\n⏳ রিভিউতে আছে: [{r}]\n❌ বাতিল হয়েছে: [{rej}]",
        "select_cat": "📋 ক্যাটাগরি নির্বাচন করুন:",
        "task_hidden": "❌ এই কাজটি বর্তমানে এডমিন দ্বারা হাইড করা আছে।",
        "no_tasks": "❌ এই ক্যাটাগরিতে বর্তমানে কোনো কাজ নেই।",
        "choose_type": "🎯 কাজ বেছে নিন:",
        "send_2fa_secret": "👉 অনুগ্রহ করে আপনার 2FA সিক্রেট কি (Secret Key) পাঠান",
        "send_cookies": "👉 অনুগ্রহ করে আপনার কুকিজ (Cookies) ডাটা পাঠান",
        "send_fb_uid": "👉 অনুগ্রহ করে আপনার ফেসবুক ইউআইডি (Facebook UID) পাঠান",
        "invalid_2fa": "❌ ভুল 2FA সিক্রেট কি! দয়া করে আবার সঠিক কি পাঠান:",
        "withdraw_dash": "💳 আপনার ব্যালেন্স ড্যাশবোর্ড\n\n💰 ব্যালেন্স: ${bal}\n💸 সর্বনিম্ন উইথড্র: $0.20\n💳 উইথড্র চার্জ: $0.03\n✅ আপনি পাবেন: ${rec}",
        "withdraw_min_err": "❌ Unsuccessful balance: Minimum $0.20 required",
        "select_meth": "💳 আপনার উইথড্র পদ্ধতি নির্বাচন করুন:",
        "send_num": "📱 অনুগ্রহ করে আপনার {method} নম্বরটি পাঠান",
        "enter_amt": "💰 অনুগ্রহ করে আপনার উইথড্র অ্যামাউন্ট লিখুন:",
        "insufficient": "❌ আপনার পর্যাপ্ত ব্যালেন্স নেই।",
        "min_amt_err": "❌ সর্বনিম্ন উইথড্র $0.20",
        "check_info": "📋 দয়া করে আপনার তথ্য যাচাই করুন\n\n📱 নম্বর: {num}\n💳 মাধ্যম: {method}\n💵 পরিমাণ: ${amt}\n✅ আপনি পাবেন: ${rec}\n\n✅ সব তথ্য ঠিক থাকলে নিচের কন্ডিশন বাটনে চাপুন। 👇",
        "btn_confirm": "✅ কনফার্ম করুন",
        "pay_pending": "⏳ আপনার পেমেন্ট পেন্ডিং অবস্থায় আছে।\n\n👨‍💼 অনুগ্রহ করে এডমিনের অনুমোদনের জন্য অপেক্ষা করুন।",
        "cookies_rec": "👉 কুকিজ পাওয়া গেছে। সামনে এগিয়ে যেতে নিচের বাটনে চাপুন।",
        "btn_acc_reg": "✅ অ্যাকাউন্ট রেজিস্টার্ড",
        "invite_check": "⚠️ আপনি কি ২ জন বন্ধুকে ইনভাইট করেছেন?\n\n❌ যদি আপনি ২ জন বন্ধুকে ইনভাইট না করে থাকেন, তবে আপনার রিপোর্টটি রিজেক্ট করা হবে।",
        "invite_check_short": "✅ নিচে কনফার্ম করে আপনার রিপোর্ট সাবমিট করুন।",
        "btn_subbed": "✅ হ্যাঁ | আমি সাবসক্রাইব করেছি",
        "thanks_msg": "✅ ধন্যবাদ! অনুগ্রহ করে আনফলো করবেন না। নিয়ম মেনে চলুন।",
        "report_received": "✅ আপনার রিপোর্টটি গ্রহণ করা হয়েছে!\n⏳ অনুগ্রহ করে ১৬–২৪ ঘণ্টা অপেক্ষা করুন।",
        "no_usernames_err": "❌ বর্তমানে কোনো ইউজার খালি নেই!",
        "force_join_msg": "📢 আমাদের বটটি ব্যবহার করতে নিচের চ্যানেলগুলোতে জয়েন করুন:",
        "not_joined_all": "❌ আপনি এখনো সবগুলো চ্যানেলে জয়েন করেননি! দয়া করে জয়েন করে আবার ভেরিফাই করুন।",
        "verify_success": "✅ এখন আমাদের বটটি ব্যবহার করতে পারবেন।\nআমাদের চ্যানেলে জয়েন হওয়ার জন্য ধন্যবাদ! ❤️",
        "support_msg": "⚠️ কোনো সমস্যা হলে অ্যাডমিনকে জানান।"
    }
}

# ============================================================
# DATABASE FUNCTIONS
# ============================================================

def _default_data():
    return {
        "users": {},
        "submissions": {},
        "withdrawals": {},
        "dynamic_tasks": {},
        "saved_usernames": [],
        "task_password": "shuvo9",
        "visibility": {"instagram_task": True, "facebook_task": True},
        "fb_dump": [],
        "ig_dump": []
    }

def _load():
    """মেমরি ক্যাশ থেকে ডেটা রিটার্ন করে; প্রথমবার Firebase থেকে fetch করে ক্যাশ বানায়"""
    global _DATA_CACHE
    if _DATA_CACHE is not None:
        return _DATA_CACHE
    remote = _firebase_get()
    d = remote if isinstance(remote, dict) else {}
    if "saved_usernames" not in d:
        d["saved_usernames"] = []
    if "dynamic_tasks" not in d:
        d["dynamic_tasks"] = {}
    if "task_password" not in d:
        d["task_password"] = "shuvo9"
    if "visibility" not in d:
        d["visibility"] = {"instagram_task": True, "facebook_task": True}
    if "users" not in d:
        d["users"] = {}
    if "submissions" not in d:
        d["submissions"] = {}
    if "withdrawals" not in d:
        d["withdrawals"] = {}
    if "fb_dump" not in d:
        d["fb_dump"] = []
    if "ig_dump" not in d:
        d["ig_dump"] = []
    _DATA_CACHE = d
    return _DATA_CACHE

def _save(data):
    """ক্যাশ আপডেট করে এবং Firebase-এ (persist) সেভ করে (redeploy/restart এ ডেটা থাকবে)"""
    global _DATA_CACHE
    _DATA_CACHE = data
    ok = _firebase_patch({
        "users": data.get("users", {}),
        "submissions": data.get("submissions", {}),
        "withdrawals": data.get("withdrawals", {}),
        "dynamic_tasks": data.get("dynamic_tasks", {}),
        "saved_usernames": data.get("saved_usernames", []),
        "task_password": data.get("task_password", "shuvo9"),
        "visibility": data.get("visibility", {"instagram_task": True, "facebook_task": True}),
        "fb_dump": data.get("fb_dump", []),
        "ig_dump": data.get("ig_dump", []),
    })
    if not ok:
        logger.error("⚠️ ডেটা Firebase-এ সেভ করা যায়নি! (network/permission সমস্যা হতে পারে)")

def get_or_create_user(user_id: int, username: str = "", referred_by: int = None):
    with _lock:
        data = _load()
        uid = str(user_id)
        if uid not in data["users"]:
            data["users"][uid] = {
                "user_id": user_id,
                "username": username,
                "balance": 0.0,
                "language": "bn", 
                "success_count": 0,
                "review_count": 0,
                "rejected_count": 0,
                "referred_by": referred_by
            }
            _save(data)
        return data["users"][uid]

def generate_profile_or_get_saved(task_category="instagram"):
    """টাস্ক ক্যাটাগরির উপর ভিত্তি করে ডাটা জেনারেট করে"""
    first_names = ["fatima", "wafaa", "ahmed", "youssef", "omar", "nour", "ali"]
    last_names = ["Zayan", "Emad", "Khan", "Ahmed", "Ali", "Hassan"]
    f_name = f"{random.choice(first_names)} {random.choice(last_names)}"
    
    if task_category == "facebook":
        return f_name, "facebook_no_username"
        
    with _lock:
        data = _load()
        if data.get("saved_usernames"):
            login_name = data["saved_usernames"].pop(0)
            _save(data)
            return f_name, login_name
            
    return None, None

# ============================================================
# CHECK JOIN FUNCTION
# ============================================================
async def is_user_joined_all(bot, user_id: int) -> bool:
    if user_id == ADMIN_ID:
        return True
    for channel in REQUIRED_CHANNELS:
        try:
            member = await bot.get_chat_member(chat_id=channel, user_id=user_id)
            if member.status in ["left", "kicked"]:
                return False
        except TelegramError:
            return False
    return True

def get_force_join_keyboard(lang: str):
    btn_ch1 = InlineKeyboardButton("📢 Range Channel", url="https://t.me/range_channele")
    btn_ch2 = InlineKeyboardButton("📢 Instagram TH", url="https://t.me/insagramth")
    btn_verify = InlineKeyboardButton("✅ Verify Membership", callback_data="verify_join")
    
    _style(btn_ch1, 'primary')
    _style(btn_ch2, 'primary')
    _style(btn_verify, 'success')
    
    return InlineKeyboardMarkup([[btn_ch1], [btn_ch2], [btn_verify]])

# ============================================================
# KEYBOARDS DEFINITION
# ============================================================

def main_menu_keyboard(user_id: int, lang: str):
    ln = LANGUAGES[lang]
    
    btn_balance = KeyboardButton(ln["btn_balance"])
    btn_tasks = KeyboardButton(ln["btn_tasks"])
    btn_withdraw = KeyboardButton(ln["btn_withdraw"])
    btn_report = KeyboardButton(ln["btn_report"])
    btn_support = KeyboardButton(ln["btn_support"])
    btn_language = KeyboardButton(ln["btn_language"])
    btn_refer = KeyboardButton("🔗 Refer" if lang == "en" else "🔗 রেফার করুন")
    
    _style(btn_balance, 'success')
    _style(btn_tasks, 'primary')
    _style(btn_withdraw, 'success')
    _style(btn_report, 'primary')
    _style(btn_support, 'primary')
    _style(btn_language, 'primary')
    _style(btn_refer, 'success')
    
    buttons = [
        [btn_balance, btn_tasks],
        [btn_withdraw, btn_report],
        [btn_support, btn_language],
        [btn_refer]
    ]
    
    if user_id == ADMIN_ID:
        btn_admin = KeyboardButton(ln["btn_admin"])
        _style(btn_admin, 'danger')
        buttons.append([btn_admin])
        
    return ReplyKeyboardMarkup(buttons, resize_keyboard=True)

# ============================================================
# USER STATE MANAGEMENT
# ============================================================
USER_STATE = {}

# ============================================================
# HANDLERS
# ============================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user

    # রেফারেল লিংক দিয়ে এসেছে কিনা চেক করা হচ্ছে (/start ref_<referrer_id>)
    referrer_id = None
    if context.args and context.args[0].startswith("ref_"):
        try:
            candidate_id = int(context.args[0].replace("ref_", ""))
            if candidate_id != user.id:  # নিজের লিংক দিয়ে নিজে এলে গণনা হবে না
                referrer_id = candidate_id
        except ValueError:
            pass

    is_new_user = str(user.id) not in _load().get("users", {})
    u_data = get_or_create_user(user.id, user.username or "", referred_by=referrer_id)
    lang = u_data.get("language", "bn")

    # নতুন ইউজার হলে এবং কারো রেফার লিংক দিয়ে এসে থাকলে, রেফারারকে জানানো হচ্ছে
    if is_new_user and referrer_id:
        try:
            await context.bot.send_message(
                chat_id=referrer_id,
                text="🎉 New User Notification\n\n👤 আপনার রেফার লিংক দিয়ে একজন নতুন ইউজার বটে যুক্ত হয়েছে!\n\n💰 Your earn 10% bonus\n✅ Your bonus system is ON — এই ইউজার কোনো রিপোর্ট সফল করলেই আপনি বোনাস পাবেন।"
            )
        except Exception:
            pass
    
    if not await is_user_joined_all(context.bot, user.id):
        await update.message.reply_text(
            LANGUAGES[lang]["force_join_msg"],
            reply_markup=get_force_join_keyboard(lang)
        )
        return

    await update.message.reply_text(
        LANGUAGES[lang]["welcome"].format(name=user.first_name),
        reply_markup=main_menu_keyboard(user.id, lang)
    )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text = update.message.text.strip()
    
    db_data = _load()
    user_profile = get_or_create_user(user_id, update.effective_user.username or "")
    lang = user_profile.get("language", "bn")
    ln = LANGUAGES[lang]

    if not await is_user_joined_all(context.bot, user_id):
        await update.message.reply_text(ln["force_join_msg"], reply_markup=get_force_join_keyboard(lang))
        return
    
    if text == "❓ How to get 2fa?":
        await update.message.reply_text("📹 How to setup 2FA Video Link:\n\nhttps://t.me/range_channele/955")
        return

    # --------------------------------------------------------
    # ADMIN FLOWS
    # --------------------------------------------------------

    if user_id == ADMIN_ID and USER_STATE.get(user_id, {}).get("step") == "admin_change_password":
        USER_STATE.pop(user_id, None)
        with _lock:
            data = _load()
            data["task_password"] = text
            _save(data)
        await update.message.reply_text(f"🔐 সফলভাবে নতুন পাসওয়ার্ড সেভ করা হয়েছে!\nবর্তমান পাসওয়ার্ড: `{text}`", parse_mode="Markdown", reply_markup=main_menu_keyboard(user_id, lang))
        return

    if user_id == ADMIN_ID and USER_STATE.get(user_id, {}).get("step") == "admin_save_username":
        USER_STATE.pop(user_id, None)
        raw_names = text.replace(",", " ").split()
        with _lock:
            data = _load()
            for r_name in raw_names:
                if r_name not in data["saved_usernames"]:
                    data["saved_usernames"].append(r_name)
            _save(data)
        await update.message.reply_text(f"✅ সফলভাবে ইউজারনেম সেভ করা হয়েছে!\nবর্তমানে মোট সেভ করা ইউজারনেম: {len(data['saved_usernames'])} টি।", reply_markup=main_menu_keyboard(user_id, lang))
        return

    if user_id == ADMIN_ID and USER_STATE.get(user_id, {}).get("step") == "add_task_button_name":
        button_name = text.strip()
        
        if not button_name:
            await update.message.reply_text("❌ বাটনের নাম খালি থাকতে পারে না!")
            return
        
        USER_STATE[user_id]["task_name"] = button_name
        USER_STATE[user_id]["step"] = "select_category_for_task"
        
        btn_ig = InlineKeyboardButton("🔷 Instagram", callback_data="adm_cat:instagram:")
        btn_fb = InlineKeyboardButton("🟩 Facebook", callback_data="adm_cat:facebook:")
        btn_cancel = InlineKeyboardButton("❌ Cancel", callback_data="cancel_add_task")
        
        _style(btn_ig, 'primary')
        _style(btn_fb, 'success')
        _style(btn_cancel, 'danger')
        
        kb = InlineKeyboardMarkup([[btn_ig, btn_fb], [btn_cancel]])
        await update.message.reply_text(f"📌 বাটন নাম: {button_name}\n\nকোন ক্যাটাগরিতে যুক্ত করবেন?", reply_markup=kb)
        return

    if user_id == ADMIN_ID and USER_STATE.get(user_id, {}).get("step") == "admin_task_name":
        if "selected_task_name" in USER_STATE[user_id]:
            try:
                price = float(text)
                USER_STATE[user_id]["task_price"] = price
                USER_STATE[user_id]["step"] = "admin_task_rules"
                await update.message.reply_text("📝 টাস্কের নিয়মাবলী (Rules) লিখুন:")
            except:
                await update.message.reply_text("❌ সঠিক সংখ্যা বা ডেসিমাল অ্যামাউন্ট দিন।")
        else:
            USER_STATE[user_id]["task_name"] = text
            USER_STATE[user_id]["step"] = "admin_task_price"
            await update.message.reply_text("💵 টাস্কের দাম কত হবে লিখুন (যেমন: 3.5):")
        return

    if user_id == ADMIN_ID and USER_STATE.get(user_id, {}).get("step") == "admin_task_price":
        try:
            price = float(text)
            USER_STATE[user_id]["task_price"] = price
            USER_STATE[user_id]["step"] = "admin_task_rules"
            await update.message.reply_text("📝 টাস্কের নিয়মাবলী (Rules) লিখুন:")
        except:
            await update.message.reply_text("❌ সঠিক সংখ্যা বা ডেসিমাল অ্যামাউন্ট দিন।")
        return

    if user_id == ADMIN_ID and USER_STATE.get(user_id, {}).get("step") == "admin_task_rules":
        USER_STATE[user_id]["task_rules"] = text
        USER_STATE[user_id]["step"] = "admin_task_type"
        
        btn_c_work = InlineKeyboardButton("🍪 Cookies Work", callback_data="adm_t_type:cookies")
        btn_2_work = InlineKeyboardButton("🛡️ 2FA Work", callback_data="adm_t_type:2fa")
        
        _style(btn_c_work, 'success')
        _style(btn_2_work, 'primary')
        
        kb = InlineKeyboardMarkup([[btn_c_work], [btn_2_work]])
        await update.message.reply_text("🎯 এটি কি ধরনের কাজ হবে নিচে থেকে সিলেক্ট করুন:", reply_markup=kb)
        return

    if user_id == ADMIN_ID and USER_STATE.get(user_id, {}).get("step") == "broadcast_msg":
        USER_STATE.pop(user_id, None)
        all_users = db_data["users"].keys()
        count = 0
        for u in all_users:
            try:
                await context.bot.send_message(chat_id=int(u), text=f"📢 ADMIN NOTICE:\n\n{text}")
                count += 1
            except:
                pass
        await update.message.reply_text(f"✅ Broadcast sent to {count} users.", reply_markup=main_menu_keyboard(user_id, lang))
        return

    if user_id == ADMIN_ID and USER_STATE.get(user_id, {}).get("step") == "add_money_uid":
        USER_STATE[user_id]["target_uid"] = text
        USER_STATE[user_id]["step"] = "add_money_amount"
        await update.message.reply_text("💵 Enter Amount to Add:")
        return
        
    if user_id == ADMIN_ID and USER_STATE.get(user_id, {}).get("step") == "add_money_amount":
        try:
            amount = float(text)
            target = USER_STATE[user_id]["target_uid"]
            with _lock:
                data = _load()
                if target in data["users"]:
                    data["users"][target]["balance"] = round(data["users"][target]["balance"] + amount, 2)
                    _save(data)
                    await update.message.reply_text(f"✅ Added ${amount} to UID {target}")
                    try:
                        await context.bot.send_message(chat_id=int(target), text=f"💰 Admin added ${amount} to your balance!")
                    except:
                        pass
                else:
                    await update.message.reply_text("❌ User not found.")
        except:
            await update.message.reply_text("❌ Invalid Amount.")
        USER_STATE.pop(user_id, None)
        return

    # --- WITHDRAW PROCESS ---
    if USER_STATE.get(user_id, {}).get("step") == "withdraw_num":
        if text == ln["btn_cancel"] or text.lower() == "cancel":
            USER_STATE.pop(user_id, None)
            await update.message.reply_text(ln["btn_cancel"], reply_markup=main_menu_keyboard(user_id, lang))
            return
        USER_STATE[user_id]["number"] = text
        USER_STATE[user_id]["step"] = "withdraw_amt"
        await update.message.reply_text(ln["enter_amt"])
        return

    if USER_STATE.get(user_id, {}).get("step") == "withdraw_amt":
        try:
            amt = float(text)
            method = USER_STATE[user_id]["method"]
            num = USER_STATE[user_id]["number"]

            if amt < 0.20:
                await update.message.reply_text("❌ Minimum withdraw is $0.20")
                USER_STATE.pop(user_id, None)
                return
            if amt > user_profile["balance"]:
                await update.message.reply_text(ln["insufficient"])
                USER_STATE.pop(user_id, None)
                return

            charge = 0.03
            receive = round(amt - charge, 4)

            USER_STATE[user_id]["amt"] = amt
            USER_STATE[user_id]["receive"] = receive

            label = "Binance UID" if method == "Binance" else "Number"
            confirm_text = (
                f"📋 Please Check Your Info\n\n"
                f"{'🟡' if method == 'Binance' else '📱'} {label}: {num}\n"
                f"💳 Method: {method}\n"
                f"💵 Amount: ${amt}\n"
                f"💳 Fee: ${charge}\n"
                f"✅ Receive: ${receive}\n\n"
                f"✅ If All Information Is Correct, Please Tap the Confirm Button. 👇"
            )
            
            btn_cnf = KeyboardButton(ln["btn_confirm"])
            btn_cnc = KeyboardButton(ln["btn_cancel"])
            _style(btn_cnf, 'success')
            _style(btn_cnc, 'danger')
            
            kb = ReplyKeyboardMarkup([[btn_cnf, btn_cnc]], resize_keyboard=True)
            USER_STATE[user_id]["step"] = "withdraw_confirm"
            await update.message.reply_text(confirm_text, reply_markup=kb)
        except:
            await update.message.reply_text("❌ Invalid format.")
            USER_STATE.pop(user_id, None)
        return

    if USER_STATE.get(user_id, {}).get("step") == "withdraw_confirm":
        if text == ln["btn_confirm"]:
            state = USER_STATE[user_id]
            w_id = str(uuid.uuid4())[:8]
            with _lock:
                data = _load()
                data["withdrawals"][w_id] = {
                    "w_id": w_id, "user_id": user_id, "username": user_profile["username"],
                    "number": state["number"], "method": state["method"], "amount": state["amt"], "status": "pending"
                }
                _save(data)
            await update.message.reply_text(ln["pay_pending"], reply_markup=main_menu_keyboard(user_id, lang))
            
            btn_w_ap = InlineKeyboardButton("✅ APPROVE", callback_data=f"w_app:{w_id}")
            btn_w_rj = InlineKeyboardButton("❌ REJECT", callback_data=f"w_rej:{w_id}")
            _style(btn_w_ap, 'success')
            _style(btn_w_rj, 'danger')
            
            admin_kb = InlineKeyboardMarkup([[btn_w_ap, btn_w_rj]])
            label = "Binance UID" if state["method"] == "Binance" else "Number"
            admin_text = (
                f"💸 New Payment Request\n\n"
                f"👤 User: @{user_profile['username']}\n🆔 UID: {user_id}\n"
                f"{'🟡' if state['method'] == 'Binance' else '📱'} {label}: {state['number']}\n"
                f"💳 Method: {state['method']}\n"
                f"💵 Amount: ${state['amt']}\n"
                f"✅ Send: ${state.get('receive', '')}"
            )
            await context.bot.send_message(
                chat_id=ADMIN_ID,
                text=admin_text,
                reply_markup=admin_kb
            )
        else:
            await update.message.reply_text(ln["btn_cancel"], reply_markup=main_menu_keyboard(user_id, lang))
        USER_STATE.pop(user_id, None)
        return

    # --- FACEBOOK UID SUBMISSION ---
    if USER_STATE.get(user_id, {}).get("step") == "waiting_for_fb_uid":
        if text == ln["btn_cancel"] or text.lower() == "cancel":
            USER_STATE.pop(user_id, None)
            await update.message.reply_text(ln["btn_cancel"], reply_markup=main_menu_keyboard(user_id, lang))
            return
        USER_STATE[user_id]["fb_uid"] = text
        USER_STATE[user_id]["step"] = "waiting_for_cookies"
        
        btn_cnc = KeyboardButton(ln["btn_cancel"])
        _style(btn_cnc, 'danger')
        await update.message.reply_text(ln["send_cookies"], reply_markup=ReplyKeyboardMarkup([[btn_cnc]], resize_keyboard=True))
        return

    # --- COOKIES SUBMISSION ---
    if USER_STATE.get(user_id, {}).get("step") == "waiting_for_cookies":
        if text == ln["btn_cancel"] or text.lower() == "cancel":
            USER_STATE.pop(user_id, None)
            await update.message.reply_text(ln["btn_cancel"], reply_markup=main_menu_keyboard(user_id, lang))
            return
        if len(text.strip()) < 100:
            await update.message.reply_text("❌ Cookie is too short (minimum 100 characters). Please provide valid cookie data.")
            return
        USER_STATE[user_id]["cookies_data"] = text
        USER_STATE[user_id]["step"] = "cookies_submitted"
        
        btn_reg = KeyboardButton(ln["btn_acc_reg"])
        btn_cnc = KeyboardButton(ln["btn_cancel"])
        _style(btn_reg, 'success')
        _style(btn_cnc, 'danger')
        
        kb = ReplyKeyboardMarkup([[btn_reg], [btn_cnc]], resize_keyboard=True)
        await update.message.reply_text(ln["cookies_rec"], reply_markup=kb)
        return

    if text == ln["btn_acc_reg"]:
        state = USER_STATE.get(user_id)
        if state and (state.get("step") == "cookies_submitted" or state.get("step") == "2fa_verify"):
            btn_sub = KeyboardButton(ln["btn_subbed"])
            btn_cnc = KeyboardButton(ln["btn_cancel"])
            _style(btn_sub, 'success')
            _style(btn_cnc, 'danger')
            
            kb = ReplyKeyboardMarkup([[btn_sub], [btn_cnc]], resize_keyboard=True)
            if state.get("step") == "cookies_submitted":
                await update.message.reply_text(ln["invite_check"], reply_markup=kb)
                USER_STATE[user_id]["step"] = "cookies_final_confirm"
            else:
                await update.message.reply_text(ln["invite_check_short"], reply_markup=kb)
                USER_STATE[user_id]["step"] = "2fa_final_confirm"
            return

    if text == ln["btn_subbed"]:
        state = USER_STATE.get(user_id)
        if state and state.get("step") == "2fa_final_confirm":
            sub_id = str(uuid.uuid4())[:8]
            file_path = f"submission_{sub_id}.txt"
            if state.get("cat") == "facebook":
                content_text = f"Dynamic 2FA Report\nTask Name: {state.get('t_name','')}\nFirst name: {state['f_name']}\nPassword: {state['pass']}\n2FA Key: {state.get('secret','')}"
            else:
                content_text = f"Dynamic 2FA Report\nTask Name: {state.get('t_name','')}\nUsername: {state['login']}\nPassword: {state['pass']}\n2FA Key: {state.get('secret','')}"
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(content_text)
                
            with _lock:
                d = _load()
                d["submissions"][sub_id] = {
                    "sub_id": sub_id, "user_id": user_id, "username": user_profile["username"],
                    "task_type": "2fa", "task_id": state.get("task_id"), "login": state['login'], "fb_uid": state.get("fb_uid", ""), "category": state.get("cat"), "status": "pending"
                }
                d["users"][str(user_id)]["review_count"] += 1
                _save(d)
                
            with open(file_path, "rb") as f:
                await context.bot.send_document(chat_id=ADMIN_ID, document=f, caption=f"实时 Dynamic 2FA Task\nUser: @{user_profile['username']}\nUID: {user_id}")
            os.remove(file_path)

            dump_entry = {
                "task": state.get("t_name", "অজানা টাস্ক"),
                "username": state.get('f_name') if state.get("cat") == "facebook" else state.get('login'),
                "password": state.get('pass', ''),
                "twofa": state.get('secret', '')
            }
            add_to_id_dump(state.get("cat", "instagram"), dump_entry)

            await update.message.reply_text(ln["report_received"], reply_markup=main_menu_keyboard(user_id, lang))
            USER_STATE.pop(user_id, None)
            return

        if state and state.get("step") == "cookies_final_confirm":
            sub_id = str(uuid.uuid4())[:8]
            file_path = f"submission_{sub_id}.txt"
            if state.get("cat") == "facebook":
                content_text = f"Task Name: {state['t_name']}\nFirst name: {state['f_name']}\nFacebook UID: {state.get('fb_uid','')}\nPassword: {state['pass']}\nCookies: {state['cookies_data']}"
            else:
                content_text = f"Task Name: {state['t_name']}\nUsername: {state['login']}\nPassword: {state['pass']}\nCookies: {state['cookies_data']}"
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(content_text)
            with _lock:
                data = _load()
                data["submissions"][sub_id] = {
                    "sub_id": sub_id, "user_id": user_id, "username": user_profile["username"],
                    "task_type": "cookies", "task_id": state.get("task_id"), "login": state['login'], "fb_uid": state.get("fb_uid", ""), "category": state.get("cat"), "status": "pending"
                }
                data["users"][str(user_id)]["review_count"] += 1
                _save(data)
            with open(file_path, "rb") as f:
                await context.bot.send_document(chat_id=ADMIN_ID, document=f, caption=f"🍪 Dynamic Cookies Task Submission\nUser: @{user_profile['username']}\nUID: {user_id}\nSub ID: {sub_id}")
            os.remove(file_path)

            dump_entry = {
                "task": state.get("t_name", "অজানা টাস্ক"),
                "username": state.get('f_name') if state.get("cat") == "facebook" else state.get('login'),
                "password": state.get('pass', ''),
                "twofa": ""
            }
            add_to_id_dump(state.get("cat", "instagram"), dump_entry)

            await update.message.reply_text(ln["report_received"], reply_markup=main_menu_keyboard(user_id, lang))
            USER_STATE.pop(user_id, None)
            return

    # --- 2FA SECRET KEY SUBMISSION ---
    if USER_STATE.get(user_id, {}).get("step") == "waiting_for_2fa":
        if text == ln["btn_cancel"] or text.lower() == "cancel":
            await update.message.reply_text(ln["btn_cancel"], reply_markup=main_menu_keyboard(user_id, lang))
            USER_STATE.pop(user_id, None)
            return
        user_secret = text.replace(" ", "").upper()

        # আসল 2FA secret key শুধু A-Z আর 2-7 অক্ষর দিয়ে তৈরি হয় (base32),
        # আর সাধারণত ১৬-৬৪ ক্যারেক্টার লম্বা হয়। এর বাইরে কিছু হলে (0,1,8,9,
        # স্পেশাল ক্যারেক্টার, বা খুব ছোট/এলোমেলো টেক্সট) সেটা আসল secret
        # হতে পারে না — তাই সাথে সাথে বাতিল করে দেওয়া হচ্ছে।
        is_valid_format = (
            16 <= len(user_secret) <= 64
            and re.fullmatch(r"[A-Z2-7]+", user_secret) is not None
        )
        if not is_valid_format:
            await update.message.reply_text(ln["invalid_2fa"])
            return
        try:
            totp = pyotp.TOTP(user_secret)
            current_code = totp.now()
        except:
            await update.message.reply_text(ln["invalid_2fa"])
            return

        state = USER_STATE.get(user_id)
        state["secret"] = user_secret
        state["step"] = "2fa_verify"
        remaining = 30 - (int(datetime.datetime.now().timestamp()) % 30)
        
        btn_reg = KeyboardButton(ln["btn_acc_reg"])
        btn_cnc = KeyboardButton(ln["btn_cancel"])
        _style(btn_reg, 'success')
        _style(btn_cnc, 'danger')
        reg_kb = ReplyKeyboardMarkup([[btn_reg], [btn_cnc]], resize_keyboard=True)
        
        await update.message.reply_text("👉 2FA Key Received. Now verify and submit using the panel below.", reply_markup=reg_kb)

        btn_ref = InlineKeyboardButton("🔄 Refresh", callback_data="refresh_2fa_code")
        _style(btn_ref, 'primary')
        inline_kb = InlineKeyboardMarkup([[btn_ref]])
        
        msg = await update.message.reply_text(
            f"╔══════════╗\n"
            f"🔑 Your 2FA Code\n"
            f"╚══════════╝\n\n"
            f"🔢 Code : `{current_code}`\n\n"
            f"⏱️ Valid : {remaining}s\n"
            f"━━━━━━━━━━",
            parse_mode="Markdown",
            reply_markup=inline_kb
        )
        state["code_msg_id"] = msg.message_id
        return

    # --- MENU NAVIGATION ---
    if text in ["💳 BALANCE", "💳 ব্যালেন্স"]:
        await update.message.reply_text(ln["balance_msg"].format(bal=user_profile['balance']), reply_markup=main_menu_keyboard(user_id, lang))
        return

    if text in ["📊 YOUR REPORT", "📊 আপনার রিপোর্ট"]:
        await update.message.reply_text(
            ln["report_msg"].format(s=user_profile.get('success_count', 0), r=user_profile.get('review_count', 0), rej=user_profile.get('rejected_count', 0)),
            reply_markup=main_menu_keyboard(user_id, lang)
        )
        return

    if text in ["ℹ️ SUPPORT", "ℹ️ সাপোর্ট (SUPPORT)"]:
        btn_adm = InlineKeyboardButton("👤 Admin", url="https://t.me/adim_shuvo")
        _style(btn_adm, 'primary')
        inline_kb = InlineKeyboardMarkup([[btn_adm]])
        await update.message.reply_text(ln["support_msg"], reply_markup=inline_kb)
        return

    if text in ["🌐 LANGUAGE", "🌐 ভাষা (LANGUAGE)"]:
        btn_bn = InlineKeyboardButton("🇧🇩 বাংলা", callback_data="lang_bn")
        btn_en = InlineKeyboardButton("🇬🇧 English", callback_data="lang_en")
        _style(btn_bn, 'success')
        _style(btn_en, 'primary')
        
        kb = InlineKeyboardMarkup([[btn_bn, btn_en]])
        await update.message.reply_text(ln["select_lang"], reply_markup=kb)
        return

    if text in ["📋 TASKS", "📋 কাজ (TASKS)"]:
        btn_ig_cat = KeyboardButton("🔥𝗜𝗡𝗦𝗧𝗔𝗚𝗥𝗔𝗠 𝗧𝗔𝗦𝗞")
        btn_fb_cat = KeyboardButton("📘𝗙𝗔𝗖𝗘𝗕𝗢𝗢𝗞 𝗧𝗔𝗦𝗞")
        btn_back = KeyboardButton(ln["btn_back"])
        
        _style(btn_ig_cat, 'primary')
        _style(btn_fb_cat, 'success')
        _style(btn_back, 'danger')
        
        vertical_keyboard = [
            [btn_ig_cat],
            [btn_fb_cat],
            [btn_back]
        ]
        
        await update.message.reply_text(
            ln["select_cat"],
            reply_markup=ReplyKeyboardMarkup(vertical_keyboard, resize_keyboard=True)
        )
        return

    if text in ["🔥𝗜𝗡𝗦𝗧𝗔𝗚𝗥𝗔𝗠 𝗧𝗔𝗦𝗞", "📘𝗙𝗔𝗖𝗘𝗕𝗢𝗢𝗞 𝗧𝗔𝗦𝗞", "🔥𝗜𝗡𝗦𝗧𝗔𝗚𝗥𝗔`𝗠 𝗧𝗔𝗦𝗞"]:
        cat_key = "instagram" if "𝗜𝗡𝗦𝗧𝗔𝗚𝗥𝗔" in text else "facebook"
        if not db_data["visibility"].get(f"{cat_key}_task", True) and user_id != ADMIN_ID:
            await update.message.reply_text(ln["task_hidden"])
            return
            
        active_tasks = [t for t in db_data.get("dynamic_tasks", {}).values() if t.get("category") == cat_key]
        if not active_tasks:
            await update.message.reply_text(ln["no_tasks"])
            return
            
        sub_tasks = []
        for t in active_tasks:
            btn_t = KeyboardButton(f"{t['name']} (${t['price']})")
            _style(btn_t, 'success')
            sub_tasks.append([btn_t])
            
        btn_cnc = KeyboardButton(ln["btn_cancel"])
        _style(btn_cnc, 'danger')
        sub_tasks.append([btn_cnc])
        await update.message.reply_text(ln["choose_type"], reply_markup=ReplyKeyboardMarkup(sub_tasks, resize_keyboard=True))
        return

    all_tasks_for_match = db_data.get("dynamic_tasks", {})
    target_task = None
    for tid, t in all_tasks_for_match.items():
        label = f"{t['name']} (${t['price']})"
        if text == label:
            target_task = t
            break

    if target_task:
        tid = target_task["id"]
        USER_STATE[user_id] = {"task_id": tid, "task_type": target_task["type"], "cat": target_task.get("category")}

        btn_str = KeyboardButton(ln["btn_start"])
        btn_vid = KeyboardButton(ln["btn_video"])
        btn_cnc = KeyboardButton(ln["btn_cancel"])

        _style(btn_str, 'success')
        _style(btn_vid, 'primary')
        _style(btn_cnc, 'danger')

        kb = ReplyKeyboardMarkup([[btn_str], [btn_vid], [btn_cnc]], resize_keyboard=True)
        rules_msg = f"🛡️ 🌟 *{target_task['name']}*\n\n💵 Payout: ${target_task['price']}\n\n📝 *Rules:*\n{target_task['rules']}\n\n🚀 Tap START to continue."
        await update.message.reply_text(rules_msg, parse_mode="Markdown", reply_markup=kb)
        return

    if text == ln["btn_video"] or text == "🎥 ভিডিও দেখুন":
        state = USER_STATE.get(user_id, {})
        task_id = state.get("task_id")
        t_data = db_data.get("dynamic_tasks", {}).get(task_id) if task_id else None
        if t_data and t_data.get("video_file_id"):
            await update.message.reply_video(video=t_data["video_file_id"], caption=f"🎥 {t_data['name']} — কীভাবে কাজটি করবেন")
        else:
            await update.message.reply_text("❌ এই টাস্কের জন্য এখনো কোনো ভিডিও যুক্ত করা হয়নি।")
        return

    if text == ln["btn_start"]:
        state = USER_STATE.get(user_id)
        if state and "task_id" in state:
            t_data = db_data["dynamic_tasks"].get(state["task_id"])
            if t_data:
                task_cat = state.get("cat", "instagram")
                f_name, login_name = generate_profile_or_get_saved(task_category=task_cat)
                
                if login_name is None and task_cat == "instagram":
                    await update.message.reply_text(ln["no_usernames_err"])
                    return
                    
                pass_val = db_data.get("task_password", "shuvo9")
                state["login"] = login_name
                state["pass"] = pass_val
                state["f_name"] = f_name
                state["t_name"] = t_data["name"]
                
                if task_cat == "facebook":
                    mono_msg = (
                        f"First name: `{f_name}`\n"
                        f"Password: `{pass_val}`"
                    )
                else:
                    mono_msg = (
                        f"First name: `{f_name}`\n"
                        f"Login: `{login_name}`\n"
                        f"Password: `{pass_val}`"
                    )
                    
                cred_msg = await update.message.reply_text(mono_msg, parse_mode="Markdown")
                asyncio.create_task(delete_message_after_delay(context, update.effective_chat.id, cred_msg.message_id, 60))
                
                if state["task_type"] == "2fa":
                    state["step"] = "waiting_for_2fa"
                    btn_2fa = KeyboardButton(ln["btn_how_to_2fa"])
                    btn_cnc = KeyboardButton(ln["btn_cancel"])
                    _style(btn_2fa, 'primary')
                    _style(btn_cnc, 'danger')
                    task_2fa_kb = ReplyKeyboardMarkup([[btn_2fa], [btn_cnc]], resize_keyboard=True)
                    await update.message.reply_text(ln["send_2fa_secret"], reply_markup=task_2fa_kb)
                else:
                    if task_cat == "facebook":
                        state["step"] = "waiting_for_fb_uid"
                        btn_cnc = KeyboardButton(ln["btn_cancel"])
                        _style(btn_cnc, 'danger')
                        await update.message.reply_text(ln["send_fb_uid"], reply_markup=ReplyKeyboardMarkup([[btn_cnc]], resize_keyboard=True))
                    else:
                        state["step"] = "waiting_for_cookies"
                        btn_cnc = KeyboardButton(ln["btn_cancel"])
                        _style(btn_cnc, 'danger')
                        await update.message.reply_text(ln["send_cookies"], reply_markup=ReplyKeyboardMarkup([[btn_cnc]], resize_keyboard=True))
            return

    if text in ["🔗 Refer", "🔗 রেফার করুন"]:
        bot_username = context.bot.username
        ref_link = f"https://t.me/{bot_username}?start=ref_{user_id}"
        if lang == "bn":
            msg = (
                f"🔗 বট থেকে রেফার করে ইনকাম করুন!\n\n"
                f"আপনার রেফার লিংক:\n{ref_link}\n\n"
                f"এই লিংক বন্ধুদের সাথে শেয়ার করুন। কেউ আপনার লিংক দিয়ে বটে জয়েন করলেই আপনি নোটিফিকেশন পাবেন, "
                f"আর সে যতবার কোনো টাস্ক রিপোর্ট সফল (approve) করবে, ততবার আপনার ব্যালেন্সে বোনাস যুক্ত হবে।"
            )
        else:
            msg = (
                f"🔗 Refer From The Bot & Earn!\n\n"
                f"Your Referral Link:\n{ref_link}\n\n"
                f"Share this link with friends. When someone joins using your link, you'll get notified, "
                f"and every time their task report gets approved, you'll earn a bonus in your balance."
            )
        await update.message.reply_text(msg)
        return

    if text in ["📤 WITHDRAW", "📤 টাকা তুলুন"]:
        bal = user_profile["balance"]
        btn_wth = InlineKeyboardButton("Withdraw", callback_data="start_withdraw")
        _style(btn_wth, 'success')
        inline_wb = InlineKeyboardMarkup([[btn_wth]])
        await update.message.reply_text(ln["withdraw_dash"].format(bal=bal, rec=max(0.0, bal - 0.03)), reply_markup=inline_wb)
        return

    # --- ADMIN CONTROL DASHBOARD PANEL ---
    if text in ["🛠️ ADMIN PANEL", "🛠️ ENDMIN PANEL", "🛠️ এডমিন প্যানেল"] and user_id == ADMIN_ID:
        btn_add_t = KeyboardButton("➕ Add Task")
        btn_del_t = KeyboardButton("❌ Delete Task")
        btn_vis_t = KeyboardButton("👁️ Task Hide/Show")
        btn_brd_t = KeyboardButton("👤 User Broadcast")
        btn_add_m = KeyboardButton("➕ Add Money")
        btn_sav_u = KeyboardButton("📥 Username Save")
        btn_all_r = KeyboardButton("🗂️ All Report")
        btn_del_u = KeyboardButton("🗑️ User Delete")
        btn_pwd_t = KeyboardButton("🔐 Password Change")
        btn_ig_file = KeyboardButton("📷 Instagram File")
        btn_fb_file = KeyboardButton("📘 Facebook File")
        btn_work_vid = KeyboardButton("🎥 Work Video")
        btn_back_m = KeyboardButton(ln["btn_back"])
        
        _style(btn_add_t, 'success')
        _style(btn_del_t, 'danger')
        _style(btn_vis_t, 'primary')
        _style(btn_brd_t, 'primary')
        _style(btn_add_m, 'success')
        _style(btn_sav_u, 'success')
        _style(btn_all_r, 'primary')
        _style(btn_del_u, 'danger')
        _style(btn_pwd_t, 'primary')
        _style(btn_ig_file, 'primary')
        _style(btn_fb_file, 'primary')
        _style(btn_work_vid, 'success')
        _style(btn_back_m, 'danger')
        
        kb = ReplyKeyboardMarkup([
            [btn_add_t, btn_del_t],
            [btn_vis_t, btn_brd_t],
            [btn_add_m, btn_sav_u],
            [btn_all_r, btn_del_u],
            [btn_pwd_t, btn_work_vid],
            [btn_ig_file, btn_fb_file],
            [btn_back_m]
        ], resize_keyboard=True)
        await update.message.reply_text("🛠️ Admin Control Dashboard", reply_markup=kb)
        return

    if user_id == ADMIN_ID and text in ["📷 Instagram File", "📘 Facebook File"]:
        category = "instagram" if text == "📷 Instagram File" else "facebook"
        key = "ig_dump" if category == "instagram" else "fb_dump"
        label = "Instagram" if category == "instagram" else "Facebook"

        with _lock:
            d = _load()
            dump_list = d.get(key, [])
            if not dump_list:
                await update.message.reply_text(f"❌ এখনো পর্যন্ত কোনো নতুন {label} সাবমিশন জমা হয়নি।")
                return

            # টাস্ক নাম অনুযায়ী গ্রুপ করা হচ্ছে
            grouped = {}
            for entry in dump_list:
                task_name = entry.get("task", "অজানা টাস্ক")
                grouped.setdefault(task_name, []).append(entry)

            # ---------- Excel (.xlsx) ফাইল তৈরি ----------
            wb = Workbook()
            ws = wb.active
            ws.title = label[:31]

            headers = ["Task Name", "Username", "Password", "2FA"]
            ws.append(headers)
            header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
            header_font = Font(name="Arial", bold=True, color="FFFFFF")
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            group_fills = [
                PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
                PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
            ]
            row_num = 2
            for g_idx, (task_name, entries) in enumerate(grouped.items()):
                fill = group_fills[g_idx % 2]
                for e in entries:
                    ws.cell(row=row_num, column=1, value=task_name).fill = fill
                    ws.cell(row=row_num, column=2, value=e.get("username", "")).fill = fill
                    ws.cell(row=row_num, column=3, value=e.get("password", "")).fill = fill
                    ws.cell(row=row_num, column=4, value=e.get("twofa", "")).fill = fill
                    for col_idx in range(1, 5):
                        ws.cell(row=row_num, column=col_idx).font = Font(name="Arial")
                    row_num += 1

            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 22
            ws.freeze_panes = "A2"

            file_path = f"{key}_{uuid.uuid4().hex[:6]}.xlsx"
            wb.save(file_path)

            # ফাইল পাঠানোর পর এই ক্যাটাগরির ডাম্প খালি করে দেওয়া হচ্ছে,
            # যাতে পরের বার শুধু নতুন সাবমিশনগুলোই থাকে
            d[key] = []
            _save(d)

        with open(file_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=f"{label}_Submissions.xlsx",
                caption=f"📄 মোট {len(dump_list)} টি {label} সাবমিশন, {len(grouped)} টি টাস্কে ভাগ করা। (এই ফাইল পাঠানোর সাথে সাথে লিস্ট খালি হয়ে গেছে, এখন থেকে নতুন সাবমিশন জমা হবে।)"
            )
        os.remove(file_path)
        return

    if user_id == ADMIN_ID and text == "🔐 Password Change":
        USER_STATE[user_id] = {"step": "admin_change_password"}
        current_pwd = db_data.get("task_password", "shuvo9")
        await update.message.reply_text(f"🔐 বর্তমান টাস্ক পাসওয়ার্ড: `{current_pwd}`\n\nনতুন যে পাসওয়ার্ডটি সেট করতে চান তা লিখে পাঠান:", parse_mode="Markdown")
        return

    if user_id == ADMIN_ID and text == "❌ Delete Task":
        btn_del_ig = InlineKeyboardButton("Instagram Tasks", callback_data="adm_del_cat:instagram")
        btn_del_fb = InlineKeyboardButton("Facebook Tasks", callback_data="adm_del_cat:facebook")
        _style(btn_del_ig, 'danger')
        _style(btn_del_fb, 'danger')
        
        kb = InlineKeyboardMarkup([[btn_del_ig, btn_del_fb]])
        await update.message.reply_text("🗑️ কোন ক্যাটাগরির কাজ ডিলিট করতে চান?", reply_markup=kb)
        return

    if user_id == ADMIN_ID and text == "🎥 Work Video":
        btn_vid_ig = InlineKeyboardButton("Instagram Tasks", callback_data="adm_vid_cat:instagram")
        btn_vid_fb = InlineKeyboardButton("Facebook Tasks", callback_data="adm_vid_cat:facebook")
        _style(btn_vid_ig, 'primary')
        _style(btn_vid_fb, 'primary')

        kb = InlineKeyboardMarkup([[btn_vid_ig, btn_vid_fb]])
        await update.message.reply_text("🎥 কোন ক্যাটাগরির টাস্কের ভিডিও সেট করতে চান?", reply_markup=kb)
        return

    if user_id == ADMIN_ID and text == "📥 Username Save":
        USER_STATE[user_id] = {"step": "admin_save_username"}
        await update.message.reply_text("📥 আপনি যে ইউজারনেমগুলো সেভ করে রাখতে চান সেগুলো পাঠান:\n(একাধিক ইউজারনেম স্পেস বা কমা দিয়ে একসাথে পাঠাতে পারেন)")
        return

    if user_id == ADMIN_ID and text == "➕ Add Task":
        USER_STATE[user_id] = {"step": "add_task_button_name"}
        await update.message.reply_text("📝 বাটনের একটি সুন্দর নাম দিন:\n\nউদাহরণ: Like 50 Posts, Share 5 Posts, Follow Account, ইত্যাদি")
        return

    if user_id == ADMIN_ID and text == "👁️ Task Hide/Show":
        v = db_data["visibility"]
        btn_h_ig = InlineKeyboardButton(f"IG Cat [{'ON' if v.get('instagram_task',True) else 'OFF'}]", callback_data="h_ig_m")
        btn_h_fb = InlineKeyboardButton(f"FB Cat [{'ON' if v.get('facebook_task',True) else 'OFF'}]", callback_data="h_fb_m")
        _style(btn_h_ig, 'primary')
        _style(btn_h_fb, 'primary')
        
        kb = InlineKeyboardMarkup([[btn_h_ig, btn_h_fb]])
        await update.message.reply_text("👁️ Click to Toggle Category Visibility:", reply_markup=kb)
        return

    if user_id == ADMIN_ID and text == "👤 User Broadcast":
        USER_STATE[user_id] = {"step": "broadcast_msg"}
        await update.message.reply_text("📢 Send the message you wish to broadcast to all users:")
        return

    if user_id == ADMIN_ID and text == "➕ Add Money":
        USER_STATE[user_id] = {"step": "add_money_uid"}
        await update.message.reply_text("👤 Send the target User identification (UID) number:")
        return

    if user_id == ADMIN_ID and text == "🗑️ User Delete":
        total_saved_usernames = len(db_data.get("saved_usernames", []))
        btn_conf_del = InlineKeyboardButton("⚠️ ডিলিট নিশ্চিত করুন", callback_data="adm_confirm_delete_all_saved_usernames")
        _style(btn_conf_del, 'danger')
        
        kb = InlineKeyboardMarkup([[btn_conf_del]])
        await update.message.reply_text(
            f"📊 বটের মোট সেভ করা ইউজারনেমের সংখ্যা: {total_saved_usernames} টি।\n\n"
            f"❗ আপনি যদি বটের সমস্ত সেভ করা ইউজারনেম ডাটা ডিলিট করতে চান তবে নিচের ডিলিট বাটনে ক্লিক করুন।",
            reply_markup=kb
        )
        return

    if user_id == ADMIN_ID and text == "🗂️ All Report":
        pending_subs = [s for s in db_data["submissions"].values() if s["status"] == "pending"]
        if not pending_subs:
            await update.message.reply_text("✅ No pending item reports.")
            return
            
        for s in pending_subs:
            btn_app = InlineKeyboardButton("Approve", callback_data=f"rep_app:{s['sub_id']}")
            btn_rej = InlineKeyboardButton("Reject", callback_data=f"rep_rej:{s['sub_id']}")
            _style(btn_app, 'success')
            _style(btn_rej, 'danger')
            inline_ap = InlineKeyboardMarkup([[btn_app, btn_rej]])
            
            u_name = s.get('username') or f"UID: {s['user_id']}"
            t_type = "Cookies" if s.get('task_type') == "cookies" else "2FA"
            
            # 🎯 ফেসবুক এবং ইনস্টাগ্রামের ক্যাটাগরি অনুযায়ী রিপোর্ট ফরম্যাট সাজানো
            if s.get('category') == "facebook":
                report_txt = (
                    f"🆔 User : @{u_name}\n"
                    f"🍪 Type: {t_type}\n"
                    f"📱 FB UID: {s.get('fb_uid', 'N/A')}\n"
                    f"⏳ Status: Pending"
                )
            else:
                report_txt = (
                    f"🆔 User : @{u_name}\n"
                    f"🍪 Type: Cookies\n"
                    f"📱 Login: {s.get('login', 'N/A')}\n"
                    f"⏳ Status: Pending"
                )
                
            await update.message.reply_text(report_txt, reply_markup=inline_ap)
        return

    if text in [ln["btn_cancel"], ln["btn_back"], "cancel", "❌ cancel", "🔙 BACK"]:
        USER_STATE.pop(user_id, None)
        await update.message.reply_text("🔙 Menu", reply_markup=main_menu_keyboard(user_id, lang))

# ============================================================
# CALLBACK QUERY PROCESSING & BACKGROUND TIMER
# ============================================================

async def delete_message_after_delay(context: ContextTypes.DEFAULT_TYPE, chat_id: int, message_id: int, delay: int):
    await asyncio.sleep(delay)
    try:
        await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
    except:
        pass

async def callback_query(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id
    data = query.data
    
    db_data = _load()
    user_profile = get_or_create_user(user_id, query.from_user.username or "")
    lang = user_profile.get("language", "bn")
    ln = LANGUAGES[lang]

    await query.answer()

    if data == "verify_join":
        if await is_user_joined_all(context.bot, user_id):
            try: await query.delete_message()
            except: pass
            
            success_msg = await context.bot.send_message(chat_id=user_id, text=ln["verify_success"])
            asyncio.create_task(delete_message_after_delay(context, user_id, success_msg.message_id, 5))
            
            await asyncio.sleep(5)
            await context.bot.send_message(
                chat_id=user_id,
                text=ln["welcome"].format(name=query.from_user.first_name),
                reply_markup=main_menu_keyboard(user_id, lang)
            )
        else:
            await context.bot.send_message(chat_id=user_id, text=ln["not_joined_all"])
        return

    if data == "refresh_2fa_code":
        state = USER_STATE.get(user_id)
        if state and state.get("step") == "2fa_verify" and "secret" in state:
            try:
                totp = pyotp.TOTP(state["secret"])
                current_code = totp.now()
                remaining = 30 - (int(datetime.datetime.now().timestamp()) % 30)
                
                btn_ref = InlineKeyboardButton("🔄 Refresh", callback_data="refresh_2fa_code")
                _style(btn_ref, 'primary')
                inline_kb = InlineKeyboardMarkup([[btn_ref]])
                
                await query.edit_message_text(
                    f"╔══════════╗\n"
                    f"🔑 Your 2FA Code\n"
                    f"╚══════════╝\n\n"
                    f"🔢 Code : `{current_code}`\n\n"
                    f"⏱️ Valid : {remaining}s\n"
                    f"━━━━━━━━━━",
                    parse_mode="Markdown",
                    reply_markup=inline_kb
                )
            except:
                pass
        return

    if data == "cancel_add_task":
        USER_STATE.pop(user_id, None)
        await query.edit_message_text("❌ বাতিল করা হয়েছে। Admin Panel 에 ফিরে গেছেন।")
        return

    if data.startswith("adm_cat:"):
        parts = data.split(":")
        cat = parts[1]
        state = USER_STATE.get(user_id, {})
        
        if state.get("step") == "select_category_for_task":
            task_name = state.get("task_name")
            USER_STATE[user_id] = {"category": cat, "step": "admin_task_price", "task_name": task_name}
            await query.message.reply_text(f"📝 বাটন নাম: {task_name}\n📁 Category: {cat}\n\n💵 টাস্কের দাম কত হবে লিখুন (যেমন: 5, 10.5):")
            try: await query.delete_message()
            except: pass
            return
        
        task_name = parts[2] if len(parts) > 2 else None
        if task_name:
            USER_STATE[user_id] = {"category": cat, "step": "admin_task_name", "selected_task_name": task_name}
            await query.message.reply_text(f"📝 Task: {task_name}\n📁 Category: {cat}\n\n💰 টাস্কের মূল্য দিন (Price in $):")
            try: await query.delete_message()
            except: pass
        else:
            USER_STATE[user_id] = {"category": cat, "step": "admin_task_name"}
            await query.message.reply_text("📝 টাস্কের একটি সুন্দর নাম (Name) দিন:")
            try: await query.delete_message()
            except: pass
        return

    if data.startswith("adm_del_cat:"):
        cat = data.split(":")[1]
        active_tasks = [t for t in db_data.get("dynamic_tasks", {}).values() if t.get("category") == cat]
        if not active_tasks:
            await query.message.reply_text("❌ এই ক্যাটাগরিতে ডিলিট করার মতো কোনো একটিভ কাজ পাওয়া যায়নি।")
            return
            
        buttons = []
        for t in active_tasks:
            btn_t_del = InlineKeyboardButton(f"🗑️ {t['name']} (${t['price']})", callback_data=f"adm_do_del:{t['id']}")
            _style(btn_t_del, 'danger')
            buttons.append([btn_t_del])
            
        await query.message.reply_text("👇 নিচে থেকে যে টাস্কটি ডিলিট করতে চান সেটির উপর চাপুন:", reply_markup=InlineKeyboardMarkup(buttons))
        try: await query.delete_message()
        except: pass
        return

    if data.startswith("adm_do_del:"):
        task_id = data.split(":")[1]
        with _lock:
            d = _load()
            if task_id in d.get("dynamic_tasks", {}):
                removed_task = d["dynamic_tasks"].pop(task_id)
                _save(d)
                await query.message.reply_text(f"✅ সফলভাবে টাস্কটি ডিলিট করা হয়েছে!\n🗑️ ডিলিট হওয়া টাস্ক: {removed_task['name']}")
            else:
                await query.message.reply_text("❌ দুঃখিত! টাস্কটি খুঁজে পাওয়া যায়নি অথবা অলরেডি ডিলিট হয়ে গেছে।")
        try: await query.delete_message()
        except: pass
        return

    if data.startswith("adm_vid_cat:"):
        cat = data.split(":")[1]
        active_tasks = [t for t in db_data.get("dynamic_tasks", {}).values() if t.get("category") == cat]
        if not active_tasks:
            await query.message.reply_text("❌ এই ক্যাটাগরিতে কোনো একটিভ টাস্ক পাওয়া যায়নি।")
            return

        buttons = []
        for t in active_tasks:
            has_vid = "🎥" if t.get("video_file_id") else "⬜"
            btn_t_vid = InlineKeyboardButton(f"{has_vid} {t['name']}", callback_data=f"adm_vid_task:{t['id']}")
            _style(btn_t_vid, 'primary')
            buttons.append([btn_t_vid])

        await query.message.reply_text("👇 কোন টাস্কের জন্য ভিডিও সেট করতে চান? (🎥 = আগে থেকেই ভিডিও সেট আছে)", reply_markup=InlineKeyboardMarkup(buttons))
        try: await query.delete_message()
        except: pass
        return

    if data.startswith("adm_vid_task:"):
        task_id = data.split(":")[1]
        t_data = db_data.get("dynamic_tasks", {}).get(task_id)
        if not t_data:
            await query.message.reply_text("❌ দুঃখিত! টাস্কটি খুঁজে পাওয়া যায়নি।")
            return
        USER_STATE[user_id] = {"step": "waiting_for_task_video", "task_id": task_id}
        await query.message.reply_text(f"🎥 '{t_data['name']}' টাস্কের জন্য এখন একটা ভিডিও পাঠান (ফাইল হিসেবে ভিডিও আপলোড করুন):")
        try: await query.delete_message()
        except: pass
        return

    if data.startswith("adm_t_type:"):
        t_type = data.split(":")[1]
        state = USER_STATE.get(user_id)
        if state and "category" in state and (state.get("selected_task_name") or state.get("task_name")):
            t_id = str(uuid.uuid4())[:8]
            new_task = {
                "id": t_id, "category": state["category"], "name": state.get("selected_task_name") or state.get("task_name"),
                "price": state.get("task_price", 0.0), "rules": state.get("task_rules", ""), "type": t_type,
                "video_file_id": None
            }
            with _lock:
                d = _load()
                if "dynamic_tasks" not in d: d["dynamic_tasks"] = {}
                d["dynamic_tasks"][t_id] = new_task
                _save(d)
                
            USER_STATE.pop(user_id, None)
            await query.message.reply_text(f"✅ সফলভাবে নতুন টাস্ক সিস্টেমে সেভ করা হয়েছে!\n\n📌 নাম: {new_task['name']}\n💵 পেমেন্ট: ${new_task['price']}\n🎯 টাইপ: {new_task['type'].upper()}")
            try: await query.delete_message()
            except: pass
        return

    if data.startswith("lang_"):
        new_lang = "bn" if data == "lang_bn" else "en"
        with _lock:
            d = _load()
            d["users"][str(user_id)]["language"] = new_lang
            _save(d)
        await query.message.reply_text(LANGUAGES[new_lang]["lang_changed"], reply_markup=main_menu_keyboard(user_id, new_lang))
        try: await query.delete_message()
        except: pass
        return

    if data == "start_withdraw":
        USER_STATE[user_id] = {"step": "withdraw_method"}
        
        btn_bks = InlineKeyboardButton("bKash", callback_data="w_meth:bKash")
        btn_ngd = InlineKeyboardButton("Nagad", callback_data="w_meth:Nagad")
        btn_bnc = InlineKeyboardButton("Binance", callback_data="w_meth:Binance")
        _style(btn_bks, 'success')
        _style(btn_ngd, 'success')
        _style(btn_bnc, 'primary')
        
        kb = InlineKeyboardMarkup([[btn_bks, btn_ngd], [btn_bnc]])
        await query.edit_message_text(LANGUAGES[lang]["select_meth"], reply_markup=kb)
        return

    if data.startswith("w_meth:"):
        method = data.split(":")[1]
        USER_STATE[user_id]["method"] = method
        USER_STATE[user_id]["step"] = "withdraw_num"
        
        btn_cnc = KeyboardButton(LANGUAGES[lang]["btn_cancel"])
        _style(btn_cnc, 'danger')
        
        if method == "Binance":
            prompt = "🟡 আপনার Binance UID দিন:" if lang == "bn" else "🟡 Please send your Binance UID:"
        else:
            prompt = LANGUAGES[lang]["send_num"].format(method=method)
        await context.bot.send_message(chat_id=user_id, text=prompt, reply_markup=ReplyKeyboardMarkup([[btn_cnc]], resize_keyboard=True))
        try: await query.delete_message()
        except: pass
        return

    if user_id != ADMIN_ID:
        return

    if data == "adm_confirm_delete_all_saved_usernames":
        with _lock:
            d = _load()
            d["saved_usernames"] = []
            _save(d)
        await query.message.reply_text("💥 সফলভাবে বটের সকল সেভ করা ইউজারনেম ডাটা ডিলিট করে দেওয়া হয়েছে!", reply_markup=main_menu_keyboard(user_id, lang))
        try: await query.delete_message()
        except: pass
        return

    if data.startswith("h_"):
        key_map = {"h_ig_m": "instagram_task", "h_fb_m": "facebook_task"}
        target_key = key_map[data]
        with _lock:
            d = _load()
            d["visibility"][target_key] = not d["visibility"].get(target_key, True)
            _save(d)
        v = d["visibility"]
        
        btn_m_ig = InlineKeyboardButton(f"IG Master [{'ON' if v.get('instagram_task',True) else 'OFF'}]", callback_data="h_ig_m")
        btn_m_fb = InlineKeyboardButton(f"FB Master [{'ON' if v.get('facebook_task',True) else 'OFF'}]", callback_data="h_fb_m")
        _style(btn_m_ig, 'primary')
        _style(btn_m_fb, 'primary')
        
        kb = InlineKeyboardMarkup([[btn_m_ig, btn_m_fb]])
        await query.edit_message_text("👁️ Category Visibility toggled:", reply_markup=kb)
        return

    if data.startswith("w_app:") or data.startswith("w_rej:"):
        w_id = data.split(":")[1]
        is_approve = data.startswith("w_app:")
        with _lock:
            d = _load()
            w_rec = d["withdrawals"].get(w_id)
            if w_rec and w_rec["status"] == "pending":
                if is_approve:
                    w_rec["status"] = "approved"
                    d["users"][str(w_rec["user_id"])]["balance"] = round(d["users"][str(w_rec["user_id"])]["balance"] - w_rec["amount"], 2)
                    msg = f"✅ Approved ${w_rec['amount']}"
                    u_msg = "✅ Your withdrawal request has been verified and approved by the admin."
                else:
                    w_rec["status"] = "rejected"
                    msg = "❌ Rejected"
                    u_msg = "❌ Your withdrawal request has been rejected."
                _save(d)
                await query.edit_message_text(msg)
                try: await context.bot.send_message(chat_id=w_rec["user_id"], text=u_msg)
                except: pass
        return

    if data.startswith("rep_app:") or data.startswith("rep_rej:"):
        sub_id = data.split(":")[1]
        is_approve = data.startswith("rep_app:")
        referrer_to_notify = None  # লকের বাইরে গিয়ে মেসেজ পাঠানোর জন্য
        with _lock:
            d = _load()
            s_rec = d["submissions"].get(sub_id)
            if s_rec and s_rec["status"] == "pending":
                u_id_str = str(s_rec["user_id"])
                if is_approve:
                    s_rec["status"] = "approved"
                    t_info = d.get("dynamic_tasks", {}).get(s_rec.get("task_id"), {})
                    p_add = t_info.get("price", 3.5)
                    d["users"][u_id_str]["balance"] = round(d["users"][u_id_str]["balance"] + p_add, 2)
                    d["users"][u_id_str]["success_count"] += 1
                    d["users"][u_id_str]["review_count"] = max(0, d["users"][u_id_str]["review_count"] - 1)
                    msg = "✅ Approved submission."
                    u_msg = f"✅ Report approved, +${p_add}"

                    # রেফারেল বোনাস — যে ইউজার কারো রেফার লিংক দিয়ে এসেছিল,
                    # তার রিপোর্ট approve হলে রেফারারকে বোনাস দেওয়া হচ্ছে
                    referrer_id = d["users"][u_id_str].get("referred_by")
                    if referrer_id and str(referrer_id) in d["users"]:
                        REFER_BONUS = 0.0005
                        d["users"][str(referrer_id)]["balance"] = round(d["users"][str(referrer_id)]["balance"] + REFER_BONUS, 4)
                        referrer_to_notify = (referrer_id, REFER_BONUS)
                else:
                    s_rec["status"] = "rejected"
                    d["users"][u_id_str]["rejected_count"] += 1
                    d["users"][u_id_str]["review_count"] = max(0, d["users"][u_id_str]["review_count"] - 1)
                    msg = "❌ Rejected submission."
                    u_msg = "❌ Your Report Has Been Rejected 🥹"
                _save(d)
                await query.edit_message_text(msg)
                try: await context.bot.send_message(chat_id=s_rec["user_id"], text=u_msg)
                except: pass
                if referrer_to_notify:
                    r_id, r_amt = referrer_to_notify
                    try:
                        await context.bot.send_message(chat_id=r_id, text=f"🎉 Your refer earn ${r_amt}")
                    except: pass

# ============================================================
# ADMIN TASK VIDEO UPLOAD HANDLER
# ============================================================
async def handle_video(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    এডমিন 'Work Video' ফিচার দিয়ে যখন কোনো টাস্কের জন্য ভিডিও আপলোড করেন,
    তখন এটা ধরে সেই টাস্কের সাথে ভিডিওটা যুক্ত করে দেয়।
    """
    user_id = update.effective_user.id
    if user_id != ADMIN_ID:
        return
    state = USER_STATE.get(user_id, {})
    if state.get("step") != "waiting_for_task_video":
        return

    task_id = state.get("task_id")
    video_file_id = update.message.video.file_id

    with _lock:
        d = _load()
        if task_id in d.get("dynamic_tasks", {}):
            d["dynamic_tasks"][task_id]["video_file_id"] = video_file_id
            _save(d)
            task_name = d["dynamic_tasks"][task_id]["name"]
            await update.message.reply_text(f"✅ '{task_name}' টাস্কের জন্য ভিডিও সফলভাবে সেট করা হয়েছে!")
        else:
            await update.message.reply_text("❌ দুঃখিত! টাস্কটি খুঁজে পাওয়া যায়নি (হয়তো ডিলিট হয়ে গেছে)।")
    USER_STATE.pop(user_id, None)

# ============================================================
# GLOBAL ERROR HANDLER
# ============================================================
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    """
    বটের যেকোনো হ্যান্ডলারে কোনো এরর হলে এটা ধরে ফেলে এবং শুধু log-এ
    সংক্ষেপে দেখায় (পুরো messy traceback console-এ আসা বন্ধ করে)।
    এটার কারণে বট কখনো ক্র্যাশ করবে না, শুধু ওই একটা রিকোয়েস্ট স্কিপ হবে।
    """
    err = context.error

    # Telegram-এর নিজস্ব Conflict এরর (একই টোকেনের একাধিক instance বা
    # deploy transition-এর সময় সাময়িকভাবে আসে) — এটা harmless, তাই
    # শুধু ছোট্ট একটা info লাইন দেখিয়ে চুপচাপ স্কিপ করা হচ্ছে।
    if isinstance(err, Conflict):
        logger.warning("Conflict: অন্য কোনো bot instance সাময়িকভাবে সক্রিয় ছিল, স্বয়ংক্রিয়ভাবে recover হচ্ছে।")
        return

    logger.error(f"হ্যান্ডলারে সমস্যা হয়েছে: {err}", exc_info=err)

    # ঠিক কোথায় (কোন ফাইলের কোন লাইনে) এররটা হয়েছে সেটা বের করা হচ্ছে,
    # যাতে এডমিনকে পাঠানো মেসেজ থেকেই বোঝা যায় সমস্যাটা কোথায় — Render
    # Logs খুলে খোঁজার দরকার না পড়ে।
    location = ""
    try:
        tb = err.__traceback__
        last_frame = tb
        while last_frame.tb_next:
            last_frame = last_frame.tb_next
        fname = os.path.basename(last_frame.tb_frame.f_code.co_filename)
        func_name = last_frame.tb_frame.f_code.co_name
        line_no = last_frame.tb_lineno
        location = f"\n\n📍 {fname} → {func_name}() → line {line_no}"
    except Exception:
        pass

    # চাইলে এডমিনকে জানিয়ে দেওয়া, কিন্তু এটাও ব্যর্থ হলে যেন বট না থামে
    try:
        if ADMIN_ID:
            await context.bot.send_message(
                chat_id=ADMIN_ID,
                text=f"⚠️ বটে একটা এরর হয়েছে (স্বয়ংক্রিয়ভাবে সামলানো হয়েছে):\n\n{type(err).__name__}: {err}{location}"
            )
    except Exception:
        pass

# ============================================================
# DUMMY HEALTH-CHECK SERVER (Render "Web Service" পোর্ট চায়)
# ============================================================
def _run_dummy_server():
    """
    Render "Web Service" একটা খোলা পোর্ট আশা করে, নাহলে সার্ভিসটাকে
    unhealthy ভেবে বারবার restart করে। বট নিজে কোনো HTTP পোর্ট ব্যবহার
    করে না বলে, শুধু Render-কে সন্তুষ্ট রাখতে একটা ছোট্ট ডামি সার্ভার
    আলাদা থ্রেডে চালানো হচ্ছে — এটা বটের আসল কাজে কোনো প্রভাব ফেলে না।
    এটা প্রোগ্রাম চলাকালীন একবারই চালু হয় (bot restart হলেও দ্বিতীয়বার না)।
    """
    port = int(os.environ.get("PORT", 10000))
    class _Health(BaseHTTPRequestHandler):
        def do_GET(self):
            self.send_response(200)
            self.end_headers()
            self.wfile.write(b"Bot is running")
        def log_message(self, *args):
            pass  # এই সার্ভারের রিকোয়েস্ট আলাদা করে log করার দরকার নেই
    try:
        HTTPServer(("0.0.0.0", port), _Health).serve_forever()
    except Exception as e:
        logger.error(f"Health-check সার্ভার চালু করা যায়নি: {e}")

# ============================================================
# MAIN EXECUTION
# ============================================================

def main():
    _load()  # স্টার্টআপেই Firebase থেকে সব ডেটা load করে ক্যাশে বসিয়ে দেয়
    logger.info("Firebase Realtime Database থেকে ডেটা সফলভাবে load হয়েছে।")
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT, handle_message))
    app.add_handler(MessageHandler(filters.VIDEO, handle_video))
    app.add_handler(CallbackQueryHandler(callback_query))
    app.add_error_handler(error_handler)
    logger.info("Bot fully updated with Custom Report Layouts.")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    import time
    threading.Thread(target=_run_dummy_server, daemon=True).start()
    # main() কোনো কারণে ক্র্যাশ করলে (যেমন নেটওয়ার্ক সাময়িকভাবে ডাউন
    # থাকলে), পুরো সার্ভিস বন্ধ না হয়ে বট নিজে থেকে কয়েক সেকেন্ড পর
    # আবার চালু হওয়ার চেষ্টা করবে।
    while True:
        try:
            main()
            break  # স্বাভাবিকভাবে থামলে (যেমন Ctrl+C) আর restart হবে না
        except Exception as e:
            logger.error(f"বট অপ্রত্যাশিতভাবে বন্ধ হয়ে গেছে, ৫ সেকেন্ড পর আবার চালু হচ্ছে: {e}", exc_info=e)
            time.sleep(5)
